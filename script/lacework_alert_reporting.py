#!/usr/bin/env python3
"""
Lacework Alert Reporting Script

This script retrieves compliance alerts from Lacework for a specified date range
and generates a comprehensive Excel report with policy details and alert information.

Features:
1. Configurable date range (defaults to previous week Mon-Sun)
2. Retrieves compliance alerts using Lacework API
3. Uses caching for policy details to avoid redundant API calls
4. Handles rate limiting with retry logic
5. Generates Excel output with comprehensive alert information and professional formatting

Output Excel fields:
- Policy ID
- Policy Title
- Description
- Remediation Steps
- Severity
- Resource (URI/workload identifier)
- Region (extracted from each URI)
- Account (AWS account ID and alias)
- Alert Status
- Alert ID (clickable hyperlink to Lacework Alert Inbox)
"""

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import laceworksdk
from tabulate import tabulate
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

try:
    from laceworksdk import LaceworkClient
except ImportError:
    print("Error: laceworksdk not installed. Please install it with: pip install laceworksdk")
    sys.exit(1)


def parse_arguments():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description='Generate Lacework compliance alert reports in Excel format',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Examples:
  # Use default (previous week Mon-Sun)
  python lacework_alert_reporting.py --api-key-file api-key/my-key.json
  
  # Specify custom date range
  python lacework_alert_reporting.py --api-key-file api-key/my-key.json --start-date 2024-01-01 --end-date 2024-01-07
  
  # Use current week Mon-Sun
  python lacework_alert_reporting.py --api-key-file api-key/my-key.json --current-week
  
  # Filter by compliance report
  python lacework_alert_reporting.py --api-key-file api-key/my-key.json --current-week -r "AWS Foundational Security Best Practices (FSBP) Standard"
  
  # Skip compliance status tab (alerts only)
  python lacework_alert_reporting.py --api-key-file api-key/my-key.json --skip-compliance
  
  # Use specific compliance report for compliance status
  python lacework_alert_reporting.py --api-key-file api-key/my-key.json --compliance-report "AWS PCI DSS 4.0.0"
        """
    )
    
    parser.add_argument(
        '-k', '--api-key-file',
        required=True,
        help='Path to the Lacework API key JSON file (e.g., api-key/my-key.json)'
    )
    
    parser.add_argument(
        '--start-date',
        help='Start date for alert retrieval (YYYY-MM-DD format)'
    )
    
    parser.add_argument(
        '--end-date',
        help='End date for alert retrieval (YYYY-MM-DD format)'
    )
    
    parser.add_argument(
        '--current-week',
        action='store_true',
        help='Use current week (Monday to Sunday) instead of previous week'
    )
    
    parser.add_argument(
        '--clear-cache',
        action='store_true',
        help='Clear all cached data before running (forces fresh API calls)'
    )
    
    parser.add_argument(
        '--output-file',
        help='Custom Excel output filename (default: auto-generated based on date range)'
    )
    
    parser.add_argument(
        '-r', '--report',
        help='Filter alerts to only include policies from the specified compliance report (e.g., "AWS Foundational Security Best Practices (FSBP) Standard")'
    )
    
    parser.add_argument(
        '--skip-compliance',
        action='store_true',
        help='Skip Compliance Status tab (only generate Alerts tab)'
    )
    
    parser.add_argument(
        '--compliance-report',
        help='Specific compliance report name to use for compliance status (e.g., "AWS Foundational Security Best Practices (FSBP) Standard")'
    )
    
    return parser.parse_args()


def get_date_range(args) -> Tuple[str, str]:
    """Determine the date range for alert retrieval."""
    if args.start_date and args.end_date:
        # Validate date format
        try:
            datetime.strptime(args.start_date, '%Y-%m-%d')
            datetime.strptime(args.end_date, '%Y-%m-%d')
            return args.start_date, args.end_date
        except ValueError:
            print("Error: Invalid date format. Use YYYY-MM-DD format.")
            sys.exit(1)
    
    # Calculate week boundaries
    today = datetime.now()
    
    if args.current_week:
        # Current week Monday to Sunday
        days_since_monday = today.weekday()
        monday = today - timedelta(days=days_since_monday)
        sunday = monday + timedelta(days=6)
    else:
        # Previous week Monday to Sunday
        days_since_monday = today.weekday()
        last_monday = today - timedelta(days=days_since_monday + 7)
        last_sunday = last_monday + timedelta(days=6)
        monday = last_monday
        sunday = last_sunday
    
    return monday.strftime('%Y-%m-%d'), sunday.strftime('%Y-%m-%d')


def load_api_credentials(api_key_file):
    """Load API credentials from JSON file."""
    # Convert to Path object for easier handling
    api_key_path = Path(api_key_file)
    
    # If no extension provided, try adding .json
    if not api_key_path.suffix:
        api_key_path = api_key_path.with_suffix('.json')
    
    # Check if file exists
    if not api_key_path.exists():
        print(f"Error: API key file not found: {api_key_path}")
        print(f"Tried: {api_key_path.absolute()}")
        sys.exit(1)
    
    try:
        with open(api_key_path, 'r') as f:
            credentials = json.load(f)
        return credentials
    except json.JSONDecodeError:
        print(f"Error: Invalid JSON in API key file: {api_key_path}")
        sys.exit(1)


def load_policy_from_cache(cache_dir: Path, policy_id: str) -> Optional[Dict]:
    """Load policy details from cache if available."""
    cache_file = cache_dir / f"{policy_id}.json"
    if cache_file.exists():
        try:
            with open(cache_file, 'r') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError) as e:
            print(f"Warning: Error reading cache file for {policy_id}: {e}")
    return None


def save_policy_to_cache(cache_dir: Path, policy_id: str, policy_data: Dict) -> None:
    """Save policy details to cache."""
    cache_file = cache_dir / f"{policy_id}.json"
    try:
        with open(cache_file, 'w') as f:
            json.dump(policy_data, f, indent=2)
    except IOError as e:
        print(f"Warning: Error writing cache file for {policy_id}: {e}")


def api_call_with_retry(func, *args, max_retries: int = 3, **kwargs):
    """Generic API call wrapper with retry logic for 429 errors."""
    for attempt in range(max_retries):
        try:
            return func(*args, **kwargs)
        
        except Exception as e:
            error_str = str(e)
            
            # Check for HTTP 429 (Too Many Requests)
            if "429" in error_str or "Too Many Requests" in error_str:
                if attempt < max_retries - 1:
                    # Try to extract Retry-After header value
                    retry_after = 60  # Default backoff
                    
                    if "Retry-After" in error_str:
                        try:
                            match = re.search(r'Retry-After[:\s]+(\d+)', error_str)
                            if match:
                                retry_after = int(match.group(1))
                        except:
                            pass
                    
                    print(f"  Rate limited (429). Waiting {retry_after} seconds before retry {attempt + 1}/{max_retries}...")
                    time.sleep(retry_after)
                    continue
            
            # For other errors or final attempt, re-raise
            if attempt == max_retries - 1:
                raise e
            else:
                print(f"  Error on attempt {attempt + 1}: {e}. Retrying...")
                time.sleep(2 ** attempt)  # Exponential backoff


def is_compliance_policy(policy_data: Dict) -> bool:
    """Check if a policy is a compliance policy."""
    policy_type = policy_data.get('policyType', '').lower()
    
    # Compliance policy types
    compliance_types = ['compliance', 'policy', 'config']
    
    # Check if policy type indicates compliance
    if policy_type in compliance_types:
        return True
    
    # Check policy title for compliance indicators
    title = policy_data.get('title', '').lower()
    compliance_keywords = ['compliance', 'config', 'security', 'audit', 'check']
    
    if any(keyword in title for keyword in compliance_keywords):
        return True
    
    return False


def get_policy_details_with_retry(client, policy_id: str, cache_dir: Path) -> Dict:
    """Retrieve policy details with caching and retry logic."""
    # First, try to load from cache
    cached_policy = load_policy_from_cache(cache_dir, policy_id)
    if cached_policy:
        return cached_policy
    
    # If not in cache, retrieve from API with retry logic
    try:
        response = api_call_with_retry(client.policies.get_by_id, policy_id)
        
        if 'data' in response:
            policy_data = response['data']
            
            # Check if this is a compliance policy
            if not is_compliance_policy(policy_data):
                print(f"  Skipping non-compliance policy: {policy_id} (type: {policy_data.get('policyType', 'Unknown')})")
                result = {
                    'policy_id': policy_id,
                    'policy_name': 'Non-Compliance Policy',
                    'severity': 'Unknown',
                    'status': 'Skipped',
                    'policy_type': policy_data.get('policyType', 'Unknown'),
                    'description': 'N/A',
                    'remediation': 'N/A',
                    'raw_data': policy_data
                }
                save_policy_to_cache(cache_dir, policy_id, result)
                return result
            
            result = {
                'policy_id': policy_id,
                'policy_name': policy_data.get('title', 'Unknown'),
                'severity': policy_data.get('severity', 'Unknown'),
                'status': 'Enabled' if policy_data.get('enabled', False) else 'Disabled',
                'policy_type': policy_data.get('policyType', 'Unknown'),
                'description': policy_data.get('description', 'N/A'),
                'remediation': policy_data.get('remediation', 'N/A'),
                'raw_data': policy_data
            }
            
            save_policy_to_cache(cache_dir, policy_id, result)
            return result
        else:
            print(f"Warning: No data returned for policy ID: {policy_id}")
            result = {
                'policy_id': policy_id,
                'policy_name': 'Not Found',
                'severity': 'Unknown',
                'status': 'Unknown',
                'policy_type': 'Unknown',
                'description': 'N/A',
                'remediation': 'N/A'
            }
            save_policy_to_cache(cache_dir, policy_id, result)
            return result
    
    except Exception as e:
        print(f"Error retrieving policy {policy_id}: {e}")
        result = {
            'policy_id': policy_id,
            'policy_name': 'Error',
            'severity': 'Unknown',
            'status': 'Unknown',
            'policy_type': 'Unknown',
            'description': 'N/A',
            'remediation': 'N/A'
        }
        save_policy_to_cache(cache_dir, policy_id, result)
        return result


def get_compliance_alerts(client, start_date: str, end_date: str) -> List[Dict]:
    """Retrieve compliance alerts for the specified date range."""
    print(f"Retrieving compliance alerts from {start_date} to {end_date}...")
    
    try:
        # Use the alerts API to get all alerts first, then filter for compliance
        response = api_call_with_retry(
            client.alerts.get,
            start_time=start_date,
            end_time=end_date
        )
        
        if 'data' in response:
            all_alerts = response['data']
            # Filter for compliance alerts only
            compliance_alerts = []
            for alert in all_alerts:
                if is_compliance_alert(alert):
                    compliance_alerts.append(alert)
            
            print(f"Found {len(compliance_alerts)} compliance alerts out of {len(all_alerts)} total alerts")
            return compliance_alerts
        else:
            print("No alerts found")
            return []
    
    except Exception as e:
        print(f"Error retrieving compliance alerts: {e}")
        return []


def get_compliance_alerts_via_cli(credentials: Dict, start_date: str, end_date: str) -> List[Dict]:
    """Get compliance alerts using Lacework CLI as fallback."""
    print(f"Retrieving compliance alerts via CLI from {start_date} to {end_date}...")
    
    try:
        # Use lacework CLI to get all alerts first, then filter for compliance
        cmd = [
            "lacework", "alert", "list",
            "--start", start_date,
            "--end", end_date,
            "--json"
        ]
        
        # Set up environment variables for Lacework CLI authentication
        env = os.environ.copy()
        env['LW_ACCOUNT'] = credentials.get('account', '')
        env['LW_API_KEY'] = credentials.get('keyId', '')
        env['LW_API_SECRET'] = credentials.get('secret', '')
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120, env=env)
        
        if result.returncode == 0:
            alerts_data = json.loads(result.stdout)
            if 'data' in alerts_data:
                all_alerts = alerts_data['data']
                # Filter for compliance alerts only
                compliance_alerts = []
                for alert in all_alerts:
                    if is_compliance_alert(alert):
                        compliance_alerts.append(alert)
                
                print(f"Found {len(compliance_alerts)} compliance alerts out of {len(all_alerts)} total alerts via CLI")
                return compliance_alerts
            else:
                print("No alerts found via CLI")
                return []
        else:
            print(f"CLI error: {result.stderr}")
            return []
    
    except subprocess.TimeoutExpired:
        print("CLI timeout")
        return []
    except json.JSONDecodeError as e:
        print(f"JSON parsing error: {e}")
        return []
    except Exception as e:
        print(f"CLI execution error: {e}")
        return []


def is_compliance_alert(alert: Dict) -> bool:
    """Check if an alert is a compliance alert based on derivedFields.sub_category."""
    # Check if this is a compliance alert using the proper field
    derived_fields = alert.get('derivedFields', {})
    sub_category = derived_fields.get('sub_category', '').lower()
    
    # Return True if sub_category is "Compliance"
    return sub_category == 'compliance'


def load_alert_from_cache(cache_dir: Path, alert_id: str) -> Optional[Dict]:
    """Load alert details from cache if available."""
    cache_file = cache_dir / f"alert_{alert_id}.json"
    if cache_file.exists():
        try:
            with open(cache_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Warning: Failed to load cached alert {alert_id}: {e}")
    return None


def save_alert_to_cache(cache_dir: Path, alert_id: str, alert_data: Dict):
    """Save alert details to cache."""
    cache_file = cache_dir / f"alert_{alert_id}.json"
    try:
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(alert_data, f, indent=2)
    except Exception as e:
        print(f"Warning: Failed to cache alert {alert_id}: {e}")


def get_detailed_alert_info(alert_id: str, cache_dir: Path, credentials: Dict) -> Dict:
    """Get detailed alert information using Lacework CLI with caching."""
    # First, try to load from cache
    cached_alert = load_alert_from_cache(cache_dir, alert_id)
    if cached_alert:
        return cached_alert
    
    # If not in cache, retrieve from CLI
    try:
        import subprocess
        
        # Set up environment variables for Lacework CLI authentication
        env = os.environ.copy()
        env['LW_ACCOUNT'] = credentials.get('account', '')
        env['LW_API_KEY'] = credentials.get('keyId', '')
        env['LW_API_SECRET'] = credentials.get('secret', '')
        
        result = subprocess.run(
            ['lacework', 'alert', 'show', str(alert_id), '--json'],
            capture_output=True,
            text=True,
            timeout=30,
            env=env
        )
        if result.returncode == 0:
            import json
            alert_data = json.loads(result.stdout)
            # Save to cache for future use
            save_alert_to_cache(cache_dir, alert_id, alert_data)
            return alert_data
        else:
            print(f"Warning: Failed to get detailed info for alert {alert_id}: {result.stderr}")
            return {}
    except Exception as e:
        print(f"Warning: Error getting detailed info for alert {alert_id}: {e}")
        return {}


def get_report_definition(report_name: str, cache_dir: Path, credentials: Dict) -> Optional[Dict]:
    """Get compliance report definition from cache or CLI with 24-hour expiration."""
    report_definitions_dir = cache_dir / "report-definitions"
    report_definitions_dir.mkdir(parents=True, exist_ok=True)
    
    cache_file = report_definitions_dir / f"{report_name.lower().replace(' ', '_').replace('/', '_')}.json"
    
    # Check if cache file exists and is less than 24 hours old
    if cache_file.exists():
        try:
            file_age = time.time() - cache_file.stat().st_mtime
            if file_age < 24 * 3600:  # 24 hours in seconds
                with open(cache_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                print(f"Report definition cache is {file_age/3600:.1f} hours old, refreshing...")
        except Exception as e:
            print(f"Warning: Failed to load cached report definition: {e}")
    
    # Get from CLI using the correct command format
    print(f"Retrieving report definition: {report_name}")
    try:
        # Set up environment variables for Lacework CLI authentication
        env = os.environ.copy()
        env['LW_ACCOUNT'] = credentials.get('account', '')
        env['LW_API_KEY'] = credentials.get('keyId', '')
        env['LW_API_SECRET'] = credentials.get('secret', '')
        
        # First, list all report definitions to find the GUID by name
        list_result = subprocess.run(
            ['lacework', 'report-definitions', 'list', '--json'],
            capture_output=True,
            text=True,
            timeout=30,
            env=env
        )
        
        if list_result.returncode != 0 or not list_result.stdout.strip():
            print(f"Warning: Failed to list report definitions: {list_result.stderr}")
            return None
        
        # Parse the list and find the GUID for the report name
        list_data = json.loads(list_result.stdout)
        report_guid = None
        
        if 'data' in list_data and isinstance(list_data['data'], list):
            for report in list_data['data']:
                if report.get('reportName', '').strip() == report_name.strip():
                    report_guid = report.get('reportDefinitionGuid')
                    break
        
        if not report_guid:
            print(f"Warning: Report definition '{report_name}' not found in available reports")
            return None
        
        # Now get the report definition using the GUID
        result = subprocess.run(
            ['lacework', 'report-definitions', 'show', report_guid, '--json'],
            capture_output=True,
            text=True,
            timeout=30,
            env=env
        )
        
        if result.returncode == 0 and result.stdout.strip():
            report_data = json.loads(result.stdout)
            # Cache the result
            try:
                with open(cache_file, 'w', encoding='utf-8') as f:
                    json.dump(report_data, f, indent=2)
                print(f"Cached report definition: {cache_file}")
            except Exception as e:
                print(f"Warning: Failed to cache report definition: {e}")
            return report_data
        else:
            print(f"Warning: Failed to get report definition '{report_name}': {result.stderr}")
            return None
    except Exception as e:
        print(f"Warning: Error getting report definition '{report_name}': {e}")
        return None


def extract_policy_ids_from_report(report_data: Dict) -> Set[str]:
    """Extract policy IDs from a compliance report definition."""
    policy_ids = set()
    
    # Handle the standard Lacework report definition format (direct data array)
    if 'data' in report_data and isinstance(report_data['data'], list):
        for section in report_data['data']:
            if 'policies' in section and isinstance(section['policies'], list):
                for policy in section['policies']:
                    if 'policyId' in policy:
                        policy_ids.add(policy['policyId'])
    
    # Handle the Lacework CLI report definition format (nested under data.reportDefinition)
    elif 'data' in report_data and 'reportDefinition' in report_data['data']:
        report_def = report_data['data']['reportDefinition']
        if 'sections' in report_def and isinstance(report_def['sections'], list):
            for section in report_def['sections']:
                if 'policies' in section and isinstance(section['policies'], list):
                    for policy_id in section['policies']:
                        if isinstance(policy_id, str):
                            policy_ids.add(policy_id)
    
    # Handle custom report definition format (direct reportDefinition)
    elif 'reportDefinition' in report_data and 'sections' in report_data['reportDefinition']:
        sections = report_data['reportDefinition']['sections']
        if isinstance(sections, list):
            for section in sections:
                if 'policies' in section and isinstance(section['policies'], list):
                    for policy_id in section['policies']:
                        if isinstance(policy_id, str):
                            policy_ids.add(policy_id)
    
    return policy_ids


def filter_alerts_by_report(alerts: List[Dict], report_name: str, cache_dir: Path, credentials: Dict) -> List[Dict]:
    """Filter alerts to only include those from policies in the specified report."""
    print(f"Filtering alerts by report: {report_name}")
    
    # Get report definition
    report_data = get_report_definition(report_name, cache_dir, credentials)
    if not report_data:
        print(f"Warning: Could not retrieve report definition for '{report_name}'. Returning all alerts.")
        return alerts
    
    # Extract policy IDs from report
    report_policy_ids = extract_policy_ids_from_report(report_data)
    print(f"Found {len(report_policy_ids)} policies in report definition")
    
    if not report_policy_ids:
        print(f"Warning: No policies found in report definition '{report_name}'. Returning all alerts.")
        return alerts
    
    # Filter alerts
    filtered_alerts = []
    for alert in alerts:
        alert_policy_id = alert.get('policyId', '')
        if alert_policy_id in report_policy_ids:
            filtered_alerts.append(alert)
    
    print(f"Filtered from {len(alerts)} to {len(filtered_alerts)} alerts matching report policies")
    return filtered_alerts


def get_aws_accounts(credentials: Dict) -> List[Dict]:
    """Get list of configured AWS accounts from Lacework."""
    print("Retrieving configured AWS accounts...")
    
    try:
        # Set up environment variables for Lacework CLI authentication
        env = os.environ.copy()
        env['LW_ACCOUNT'] = credentials.get('account', '')
        env['LW_API_KEY'] = credentials.get('keyId', '')
        env['LW_API_SECRET'] = credentials.get('secret', '')
        
        cmd = [
            "lacework", "compliance", "aws", "list-accounts", 
            "--json",
            "--api_key", credentials.get('keyId', ''),
            "--api_secret", credentials.get('secret', ''),
            "--account", credentials.get('account', '')
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60, env=env)
        
        if result.returncode == 0:
            accounts_data = json.loads(result.stdout)
            
            # Handle the actual response format from lacework compliance aws list-accounts
            if 'aws_accounts' in accounts_data and isinstance(accounts_data['aws_accounts'], list):
                accounts = []
                for account in accounts_data['aws_accounts']:
                    accounts.append({
                        'account_id': account.get('account_id', ''),
                        'account_alias': account.get('account_alias', ''),
                        'status': account.get('status', ''),
                        'integration_guid': account.get('integration_guid', '')
                    })
                
                # Filter to only enabled accounts for compliance reporting
                enabled_accounts = [acc for acc in accounts if acc['status'] == 'Enabled']
                print(f"Found {len(accounts)} total AWS accounts ({len(enabled_accounts)} enabled)")
                return enabled_accounts
            else:
                print(f"No AWS accounts found in response. Data structure: {accounts_data}")
                return []
        else:
            print(f"CLI error retrieving AWS accounts: {result.stderr}")
            return []
    
    except subprocess.TimeoutExpired:
        print("CLI timeout retrieving AWS accounts")
        return []
    except json.JSONDecodeError as e:
        print(f"JSON parsing error retrieving AWS accounts: {e}")
        return []
    except Exception as e:
        print(f"Error retrieving AWS accounts: {e}")
        return []


def get_compliance_report(account_id: str, credentials: Dict, cache_dir: Path, report_name: str = None) -> List[Dict]:
    """Get compliance report for a specific AWS account with non-compliant resources."""
    
    # Create cache directory for compliance reports
    compliance_cache_dir = cache_dir / "compliance-reports"
    compliance_cache_dir.mkdir(parents=True, exist_ok=True)
    
    # Create cache filename that includes both account ID and report name
    if report_name:
        report_suffix = report_name.lower().replace(' ', '_').replace('/', '_').replace('(', '').replace(')', '').replace(':', '').replace('.', '')
        cache_file = compliance_cache_dir / f"compliance_{account_id}_{report_suffix}.json"
    else:
        cache_file = compliance_cache_dir / f"compliance_{account_id}_default.json"
    
    # Check if cache file exists and is less than 24 hours old
    if cache_file.exists():
        try:
            file_age = time.time() - cache_file.stat().st_mtime
            if file_age < 24 * 3600:  # 24 hours in seconds
                print(f"Loading compliance report from cache (age: {file_age/3600:.1f} hours)")
                with open(cache_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                print(f"Compliance report cache is {file_age/3600:.1f} hours old, refreshing...")
        except Exception as e:
            print(f"Warning: Failed to load cached compliance report: {e}")
    
    try:
        # Set up environment variables for Lacework CLI authentication
        env = os.environ.copy()
        env['LW_ACCOUNT'] = credentials.get('account', '')
        env['LW_API_KEY'] = credentials.get('keyId', '')
        env['LW_API_SECRET'] = credentials.get('secret', '')
        
        # Build command to get compliance report
        cmd = [
            "lacework", "compliance", "aws", "get-report", account_id,
            "--status", "non-compliant",
            "--json",
            "--api_key", credentials.get('keyId', ''),
            "--api_secret", credentials.get('secret', ''),
            "--account", credentials.get('account', '')
        ]
        
        # Add report name filter if specified
        if report_name:
            cmd.extend(["--report_name", report_name])
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=300, env=env)
        
        if result.returncode == 0:
            compliance_data = json.loads(result.stdout)
            
            # Cache the result
            try:
                with open(cache_file, 'w', encoding='utf-8') as f:
                    json.dump(compliance_data, f, indent=2)
                print(f"Cached compliance report: {cache_file}")
            except Exception as e:
                print(f"Warning: Failed to cache compliance report: {e}")
            
            return compliance_data
        else:
            print(f"CLI error retrieving compliance report: {result.stderr}")
            return []
    
    except subprocess.TimeoutExpired:
        print("CLI timeout retrieving compliance report")
        return []
    except json.JSONDecodeError as e:
        print(f"JSON parsing error retrieving compliance report: {e}")
        return []
    except Exception as e:
        print(f"Error retrieving compliance report: {e}")
        return []


def parse_compliance_report_data(compliance_data: Dict, policy_details: Dict) -> List[Dict]:
    """Parse compliance report data and extract relevant fields for Excel output."""
    compliance_items = []
    
    # Handle the actual Lacework compliance report format
    if 'recommendations' in compliance_data and isinstance(compliance_data['recommendations'], list):
        items = compliance_data['recommendations']
        print(f"Processing {len(items)} compliance recommendations")
    elif 'data' in compliance_data:
        report_data = compliance_data['data']
        
        # Handle different compliance report formats
        if isinstance(report_data, list):
            # Direct list of compliance items
            items = report_data
        elif isinstance(report_data, dict) and 'recommendations' in report_data:
            # Format with recommendations array
            items = report_data['recommendations']
        elif isinstance(report_data, dict) and 'violations' in report_data:
            # Format with violations array
            items = report_data['violations']
        else:
            print(f"Unknown compliance report format: {type(report_data)}")
            return compliance_items
        
        print(f"Processing {len(items)} compliance items")
    else:
        print("No recommendations or data found in compliance report")
        return compliance_items
    
    for item in items:
        # Extract policy information
        policy_id = item.get('policyId', item.get('REC_ID', 'Unknown'))
        policy_info = policy_details.get(policy_id, {})
        
        # Skip if policy is not a compliance policy
        if policy_info.get('status') == 'Skipped':
            continue
        
        # Extract account information
        account_id = item.get('accountId', item.get('ACCOUNT_ID', 'Unknown'))
        account_alias = item.get('accountAlias', item.get('ACCOUNT_ALIAS', ''))
        
        # Format account information
        if account_alias:
            account = f"{account_id} ({account_alias})"
        else:
            account = account_id
        
        # Extract compliance status
        status = item.get('status', item.get('STATUS', 'Non-Compliant'))
        
        # Handle violations array - each violation is a separate compliance item
        violations = item.get('VIOLATIONS', [])
        if violations:
            for violation in violations:
                # Extract resource information from violation
                resource_arn = violation.get('resource', 'Unknown')
                region = violation.get('region', 'Unknown')
                
                # Format resource with account information
                if resource_arn != 'Unknown' and resource_arn.startswith('arn:aws:'):
                    if account_alias:
                        resource_arn += f" (Account: {account_id}, Alias: {account_alias})"
                    elif account_id != 'Unknown':
                        resource_arn += f" (Account: {account_id})"
                
                # Use compliance report severity (numeric) over policy cache severity (text)
                compliance_severity = item.get('SEVERITY', 'Unknown')
                if compliance_severity != 'Unknown':
                    severity = str(compliance_severity)
                else:
                    severity = policy_info.get('severity', 'Unknown')
                
                compliance_item = {
                    'policy_id': policy_id,
                    'policy_name': policy_info.get('policy_name', item.get('TITLE', 'Unknown')),
                    'description': policy_info.get('description', item.get('TITLE', 'N/A')),
                    'remediation': policy_info.get('remediation', 'N/A'),
                    'severity': severity,
                    'resource': resource_arn,
                    'region': region,
                    'account': account,
                    'compliance_status': status,
                    'raw_data': item
                }
                
                compliance_items.append(compliance_item)
        else:
            # No violations, but still create a compliance item for the policy
            compliance_item = {
                'policy_id': policy_id,
                'policy_name': policy_info.get('policy_name', item.get('TITLE', 'Unknown')),
                'description': policy_info.get('description', item.get('TITLE', 'N/A')),
                'remediation': policy_info.get('remediation', 'N/A'),
                'severity': policy_info.get('severity', str(item.get('SEVERITY', 'Unknown'))),
                'resource': account_id,  # Use account ID as resource if no specific resource
                'region': 'Unknown',
                'account': account,
                'compliance_status': status,
                'raw_data': item
            }
            
            compliance_items.append(compliance_item)
    
    print(f"Processed {len(compliance_items)} compliance items")
    return compliance_items


def extract_alert_details(alert: Dict, cache_dir: Path, credentials: Dict) -> Dict:
    """Extract relevant details from an alert."""
    # Extract basic alert information
    alert_id = alert.get('alertId', 'Unknown')
    severity = alert.get('severity', 'Unknown')
    start_time = alert.get('startTime', 'Unknown')
    
    # Extract policy information
    policy_id = 'Unknown'
    if 'policyId' in alert:
        policy_id = alert['policyId']
    elif 'data' in alert and 'policyId' in alert['data']:
        policy_id = alert['data']['policyId']
    elif 'data' in alert and 'REC_ID' in alert['data']:
        policy_id = alert['data']['REC_ID']
    
    # Get detailed alert information using CLI with caching
    detailed_alert = get_detailed_alert_info(alert_id, cache_dir, credentials)
    
    # Extract resource information from detailed alert
    resources = []
    regions = []
    accounts = []
    
    if 'entityMap' in detailed_alert and 'Resource' in detailed_alert['entityMap']:
        resource_list = detailed_alert['entityMap']['Resource']
        if isinstance(resource_list, list) and len(resource_list) > 0:
            # Extract all resources and their details
            for resource_item in resource_list:
                if 'KEY' in resource_item:
                    key_data = resource_item['KEY']
                    resource_arn = key_data.get('resource', 'Unknown')
                    if resource_arn != 'Unknown':
                        # For all AWS resources, add account and alias information
                        if resource_arn.startswith('arn:aws:'):
                            account_id = key_data.get('account_id', '')
                            account_alias = key_data.get('account_alias', '')
                            if account_alias:
                                resource_arn += f" (Account: {account_id}, Alias: {account_alias})"
                                accounts.append(f"{account_id} ({account_alias})")
                            elif account_id:
                                resource_arn += f" (Account: {account_id})"
                                accounts.append(account_id)
                        resources.append(resource_arn)
                    
                    # Extract region for each resource
                    resource_region = key_data.get('resource_region', 'Unknown')
                    if resource_region != 'Unknown':
                        regions.append(resource_region)
    
    # Fallback to data array if entityMap extraction failed
    elif 'data' in detailed_alert and isinstance(detailed_alert['data'], list) and len(detailed_alert['data']) > 0:
        for data_item in detailed_alert['data']:
            if 'KEY' in data_item:
                key_data = data_item['KEY']
                resource_arn = key_data.get('resource', 'Unknown')
                if resource_arn != 'Unknown':
                    # For all AWS resources, add account and alias information
                    if resource_arn.startswith('arn:aws:'):
                        account_id = key_data.get('account_id', '')
                        account_alias = key_data.get('account_alias', '')
                        if account_alias:
                            resource_arn += f" (Account: {account_id}, Alias: {account_alias})"
                            accounts.append(f"{account_id} ({account_alias})")
                        elif account_id:
                            resource_arn += f" (Account: {account_id})"
                            accounts.append(account_id)
                    resources.append(resource_arn)
                
                # Extract region for each resource
                resource_region = key_data.get('resource_region', 'Unknown')
                if resource_region != 'Unknown':
                    regions.append(resource_region)
    
    # Fallback to top-level fields if both extractions failed
    if not resources and 'resource' in detailed_alert:
        resources = [detailed_alert['resource']]
    if not regions and 'region' in detailed_alert:
        regions = [detailed_alert['region']]
    if not accounts and 'account' in detailed_alert:
        accounts = [detailed_alert['account']]
    
    # Format resources as a single coherent text block (like remediation steps)
    if resources:
        if len(resources) == 1:
            resource = resources[0]
        else:
            # Format as a numbered list to make it look like natural text
            resource = '\n'.join(f"{i+1}. {res}" for i, res in enumerate(resources))
    else:
        resource = 'Unknown'
    
    # Keep regions and accounts as simple joins for now
    region = '\n'.join(regions) if regions else 'Unknown'
    account = '\n'.join(accounts) if accounts else 'Unknown'
    
    # Extract source information from original alert
    derived_fields = alert.get('derivedFields', {})
    source = derived_fields.get('source', 'Unknown')
    
    return {
        'alert_id': alert_id,
        'policy_id': policy_id,
        'severity': severity,
        'start_time': start_time,
        'resource': resource,
        'region': region,
        'account': account,
        'source': source,
        'raw_alert': alert
    }


def write_alert_excel(alerts_data: List[Dict], output_file: Path, start_date: str, end_date: str, account: str, compliance_data: List[Dict] = None, report_name: str = None, compliance_report_name: str = None):
    """Write alert details and compliance status to Excel file with proper formatting."""
    alert_fieldnames = [
        'Policy ID',
        'Policy Title',
        'Description',
        'Remediation Steps',
        'Severity',
        'Resource',
        'Region',
        'Account',
        'Alert Status',
        'Alert ID'
    ]
    
    compliance_fieldnames = [
        'Policy ID',
        'Policy Title',
        'Description',
        'Remediation Steps',
        'Severity',
        'Resource',
        'Region',
        'Account'
    ]
    
    # Define severity order for sorting
    severity_order = {
        'critical': 1, 'high': 2, 'medium': 3, 'low': 4, 'info': 5, 'unknown': 6
    }
    
    # Sort alerts by severity, then source, then policy ID
    def sort_key(alert):
        severity = alert.get('severity', 'Unknown').lower()
        severity_rank = severity_order.get(severity, 999)
        source = alert.get('source', 'Unknown')
        policy_id = alert.get('policy_id', '')
        return (severity_rank, source, policy_id)
    
    sorted_alerts = sorted(alerts_data, key=sort_key)
    
    # Sort compliance data by severity, then policy ID
    def compliance_sort_key(item):
        severity = item.get('severity', 'Unknown').lower()
        severity_rank = severity_order.get(severity, 999)
        policy_id = item.get('policy_id', '')
        return (severity_rank, policy_id)
    
    sorted_compliance = sorted(compliance_data, key=compliance_sort_key) if compliance_data else []
    
    try:
        # Create workbook
        wb = Workbook()
        
        # Create Compliance Status worksheet first
        ws_compliance = wb.active
        ws_compliance.title = "Compliance Status"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # No severity colors - keep it clean and professional
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write Compliance Status header row (if compliance data exists)
        if sorted_compliance:
            for col, field in enumerate(compliance_fieldnames, 1):
                cell = ws_compliance.cell(row=1, column=col, value=field)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Write Compliance Status data rows
            for row, item in enumerate(sorted_compliance, 2):
                # Clean fields to ensure they're strings and handle None values
                def clean_field(field_value):
                    if field_value is None:
                        return 'N/A'
                    return str(field_value) if field_value else 'N/A'
                
                row_data = {
                    'Policy ID': clean_field(item['policy_id']),
                    'Policy Title': clean_field(item['policy_name']),
                    'Description': clean_field(item['description']),
                    'Remediation Steps': clean_field(item['remediation']),
                    'Severity': item['severity'].title() if item['severity'] != 'Unknown' else 'Unknown',
                    'Resource': clean_field(item['resource']),
                    'Region': clean_field(item['region']),
                    'Account': clean_field(item['account'])
                }
                
                # Write each field to the row
                for col, field in enumerate(compliance_fieldnames, 1):
                    cell = ws_compliance.cell(row=row, column=col, value=row_data[field])
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Auto-adjust Compliance Status column widths
            for col in range(1, len(compliance_fieldnames) + 1):
                column_letter = get_column_letter(col)
                
                # Calculate max width based on content
                max_length = 0
                for row in range(1, len(sorted_compliance) + 2):
                    cell_value = ws_compliance.cell(row=row, column=col).value
                    if cell_value:
                        # Count characters, but limit very long lines
                        line_lengths = [len(line) for line in str(cell_value).split('\n')]
                        max_length = max(max_length, max(line_lengths))
                
                # Set column width with reasonable limits
                adjusted_width = min(max(max_length + 2, 15), 80)
                ws_compliance.column_dimensions[column_letter].width = adjusted_width
            
            # Set Compliance Status row heights for better readability
            for row in range(2, len(sorted_compliance) + 2):
                ws_compliance.row_dimensions[row].height = 60  # Allow for wrapped text
        
        # Create Alerts worksheet
        ws_alerts = wb.create_sheet("Compliance Alerts")
        
        # Write Alerts header row
        for col, field in enumerate(alert_fieldnames, 1):
            cell = ws_alerts.cell(row=1, column=col, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write Alerts data rows
        for row, alert in enumerate(sorted_alerts, 2):
            # Clean fields to ensure they're strings and handle None values
            def clean_field(field_value):
                if field_value is None:
                    return 'N/A'
                return str(field_value) if field_value else 'N/A'
            
            # Prepare row data with cleaned fields
            alert_id = clean_field(alert['alert_id'])
            alert_url = f"https://{account}/ui/investigation/monitor/AlertInbox/{alert_id}"
            
            row_data = {
                'Policy ID': clean_field(alert['policy_id']),
                'Policy Title': clean_field(alert['policy_name']),
                'Description': clean_field(alert['description']),
                'Remediation Steps': clean_field(alert['remediation']),
                'Severity': alert['severity'].title() if alert['severity'] != 'Unknown' else 'Unknown',
                'Resource': clean_field(alert['resource']),
                'Region': clean_field(alert['region']),
                'Account': clean_field(alert['account']),
                'Alert Status': clean_field(alert['alert_status']),
                'Alert ID': alert_id
            }
            
            # Write each field to the row
            for col, field in enumerate(alert_fieldnames, 1):
                cell = ws_alerts.cell(row=row, column=col, value=row_data[field])
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Add hyperlink for Alert ID field
                if field == 'Alert ID':
                    cell.hyperlink = Hyperlink(ref=alert_url, target=alert_url)
                    cell.font = Font(color="0000FF", underline="single")
        
        # Auto-adjust Alerts column widths
        for col in range(1, len(alert_fieldnames) + 1):
            column_letter = get_column_letter(col)
            
            # Calculate max width based on content
            max_length = 0
            for row in range(1, len(sorted_alerts) + 2):
                cell_value = ws_alerts.cell(row=row, column=col).value
                if cell_value:
                    # Count characters, but limit very long lines
                    line_lengths = [len(line) for line in str(cell_value).split('\n')]
                    max_length = max(max_length, max(line_lengths))
            
            # Set column width with reasonable limits
            adjusted_width = min(max(max_length + 2, 15), 80)
            ws_alerts.column_dimensions[column_letter].width = adjusted_width
        
        # Set Alerts row heights for better readability
        for row in range(2, len(sorted_alerts) + 2):
            ws_alerts.row_dimensions[row].height = 60  # Allow for wrapped text
        
        
        # Add a summary sheet
        summary_ws = wb.create_sheet("Compliance Summary")
        
        # Determine compliance framework name
        compliance_framework = "General Compliance"
        if report_name:
            compliance_framework = report_name
        elif compliance_report_name:
            compliance_framework = compliance_report_name
        
        # Add summary information
        summary_data = [
            ["Lacework Compliance Report Summary", ""],
            ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Compliance Framework", compliance_framework],
            ["Date Range", f"{start_date} to {end_date}"],
            ["", ""],
            ["Compliance Policies (Non-Compliant)", ""]
        ]
        
        # Calculate compliance severity distribution first (if compliance data exists)
        if sorted_compliance:
            compliance_severity_counts = {}
            for item in sorted_compliance:
                severity = item['severity']  # Already mapped to text labels
                compliance_severity_counts[severity] = compliance_severity_counts.get(severity, 0) + 1
            
            for severity, count in sorted(compliance_severity_counts.items(), key=lambda x: severity_order.get(x[0], 999)):
                summary_data.append([severity, count])
        
        # Add compliance alerts section
        summary_data.extend([
            ["", ""],
            ["Compliance Alerts", ""]
        ])
        
        # Calculate alert severity distribution
        alert_severity_counts = {}
        for alert in sorted_alerts:
            severity = alert['severity'].title() if alert['severity'] != 'Unknown' else 'Unknown'
            alert_severity_counts[severity] = alert_severity_counts.get(severity, 0) + 1
        
        for severity, count in sorted(alert_severity_counts.items(), key=lambda x: severity_order.get(x[0], 999)):
            summary_data.append([severity, count])
        
        # Write summary data
        for row, (label, value) in enumerate(summary_data, 1):
            summary_ws.cell(row=row, column=1, value=label)
            summary_ws.cell(row=row, column=2, value=value)
            
            if label == "Lacework Compliance Report Summary":
                summary_ws.cell(row=row, column=1).font = Font(bold=True, size=14)
            elif label in ["Report Generated", "Compliance Framework", "Date Range", "Compliance Policies (Non-Compliant)", "Compliance Alerts"]:
                summary_ws.cell(row=row, column=1).font = Font(bold=True)
        
        # Auto-adjust summary column widths
        summary_ws.column_dimensions['A'].width = 25
        summary_ws.column_dimensions['B'].width = 20
        
        # Save the workbook
        wb.save(output_file)
        print(f"Successfully wrote {len(alerts_data)} alerts to {output_file}")
        if sorted_compliance:
            print(f"Successfully wrote {len(sorted_compliance)} compliance items to Compliance Status tab")
        
    except Exception as e:
        print(f"Error writing Excel file: {e}")
        sys.exit(1)


def clear_cache(cache_dir: Path):
    """Clear all cached data."""
    if cache_dir.exists():
        print(f"Clearing cache directory: {cache_dir}")
        shutil.rmtree(cache_dir)
        print("Cache cleared successfully")
    else:
        print("Cache directory does not exist, nothing to clear")


def display_alert_table(alerts: List[Dict]):
    """Display a table of alerts with key information."""
    if not alerts:
        print("No alerts to display")
        return
    
    # Define severity order for sorting
    severity_order = {
        'critical': 1, 'high': 2, 'medium': 3, 'low': 4, 'info': 5, 'unknown': 6
    }
    
    # Sort alerts by severity, then source, then alert name
    def sort_key(alert):
        raw_alert = alert.get('raw_alert', {})
        severity = alert.get('severity', 'Unknown').lower()
        severity_rank = severity_order.get(severity, 999)
        source = raw_alert.get('derivedFields', {}).get('source', 'Unknown')
        alert_name = raw_alert.get('alertName', raw_alert.get('title', 'Unknown'))
        return (severity_rank, source, alert_name)
    
    sorted_alerts = sorted(alerts, key=sort_key)
    
    # Prepare table data
    table_data = []
    for alert in sorted_alerts:
        raw_alert = alert.get('raw_alert', {})
        
        # Extract alert information
        alert_id = alert['alert_id']
        alert_name = raw_alert.get('alertName', raw_alert.get('title', 'Unknown'))
        alert_type = raw_alert.get('alertType', 'Unknown')
        severity = alert['severity'].title() if alert['severity'] != 'Unknown' else 'Unknown'
        status = raw_alert.get('status', 'Unknown')
        
        # Extract derived fields
        derived_fields = raw_alert.get('derivedFields', {})
        category = derived_fields.get('category', 'Unknown')
        source = derived_fields.get('source', 'Unknown')
        sub_category = derived_fields.get('sub_category', 'Unknown')
        
        # Truncate long names
        alert_name = alert_name[:35] + '...' if len(alert_name) > 35 else alert_name
        alert_type = alert_type[:20] + '...' if len(alert_type) > 20 else alert_type
        
        table_data.append([
            alert_id,
            alert_name,
            severity,
            alert_type,
            status,
            category,
            sub_category,
            source
        ])
    
    # Display table
    headers = ['Alert ID', 'Alert Name', 'Severity', 'Alert Type', 'Status', 'Category', 'Sub-Category', 'Source']
    print("\n=== Alert Summary Table ===")
    print(tabulate(table_data, headers=headers, tablefmt='grid', maxcolwidths=[10, 35, 8, 20, 8, 10, 12, 8]))
    print()


def main():
    """Main function to execute the alert reporting workflow."""
    
    # Parse command-line arguments
    args = parse_arguments()
    
    print("=== Lacework Alert Reporting Tool ===")
    print(f"API Key: {args.api_key_file}")
    
    # Determine date range
    start_date, end_date = get_date_range(args)
    print(f"Date Range: {start_date} to {end_date}")
    
    if args.clear_cache:
        print("Clear Cache: Enabled")
    print()
    
    # Define paths
    script_dir = Path(__file__).parent
    project_root = script_dir.parent
    api_key_file = Path(args.api_key_file)
    cache_dir = project_root / "cache"
    policy_cache_dir = cache_dir / "policy-details"
    alert_cache_dir = cache_dir / "alert-details"
    output_dir = project_root / "output"
    
    # Clear cache if requested
    if args.clear_cache:
        clear_cache(cache_dir)
    
    # Ensure directories exist
    policy_cache_dir.mkdir(parents=True, exist_ok=True)
    alert_cache_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Load API credentials
    print("Step 1: Loading API credentials...")
    credentials = load_api_credentials(api_key_file)
    
    # Initialize Lacework client
    print("Step 2: Initializing Lacework client...")
    try:
        client = LaceworkClient(
            account=credentials['account'],
            api_key=credentials['keyId'],
            api_secret=credentials['secret']
        )
    except Exception as e:
        print(f"Error initializing Lacework client: {e}")
        sys.exit(1)
    
    # Get compliance alerts
    print("Step 3: Retrieving compliance alerts...")
    alerts = get_compliance_alerts(client, start_date, end_date)
    
    # If no alerts via API, try CLI
    if not alerts:
        print("No alerts found via API, trying CLI...")
        alerts = get_compliance_alerts_via_cli(credentials, start_date, end_date)
    
    if not alerts:
        print("No compliance alerts found for the specified date range.")
        sys.exit(0)
    
    # Apply report filtering if specified
    if args.report:
        alerts = filter_alerts_by_report(alerts, args.report, cache_dir, credentials)
        if not alerts:
            print(f"No alerts found matching report '{args.report}'.")
            sys.exit(0)
    
    # Extract alert details
    print("Step 4: Processing alert details...")
    alert_details = []
    unique_policy_ids = set()
    
    # Check cache status for alerts
    cached_alert_count = sum(1 for alert in alerts if (alert_cache_dir / f"alert_{alert['alertId']}.json").exists())
    print(f"Found {cached_alert_count} alerts already cached, {len(alerts) - cached_alert_count} to retrieve from CLI")
    
    for alert in alerts:
        details = extract_alert_details(alert, alert_cache_dir, credentials)
        alert_details.append(details)
        if details['policy_id'] != 'Unknown':
            unique_policy_ids.add(details['policy_id'])
    
    print(f"Found {len(alert_details)} alerts with {len(unique_policy_ids)} unique policies")
    
    # Display alert table for troubleshooting
    display_alert_table(alert_details)
    
    # Get policy details for all unique policies
    print("Step 5: Retrieving policy details...")
    cached_count = sum(1 for policy_id in unique_policy_ids if (policy_cache_dir / f"{policy_id}.json").exists())
    print(f"Found {cached_count} policies already cached, {len(unique_policy_ids) - cached_count} to retrieve from API")
    
    policy_details = {}
    for policy_id in unique_policy_ids:
        policy_info = get_policy_details_with_retry(client, policy_id, policy_cache_dir)
        policy_details[policy_id] = policy_info
    
    # Enrich alert details with policy information
    print("Step 6: Enriching alerts with policy details...")
    enriched_alerts = []
    skipped_count = 0
    
    for alert in alert_details:
        policy_id = alert['policy_id']
        policy_info = policy_details.get(policy_id, {})
        
        # Skip alerts with non-compliance policies
        if policy_info.get('status') == 'Skipped':
            skipped_count += 1
            continue
        
        enriched_alert = {
            'alert_id': alert['alert_id'],
            'policy_id': policy_id,
            'policy_name': policy_info.get('policy_name', 'Unknown'),
            'description': policy_info.get('description', 'N/A'),
            'remediation': policy_info.get('remediation', 'N/A'),
            'severity': alert['severity'],
            'source': alert['source'],
            'resource': alert['resource'],
            'region': alert['region'],
            'account': alert['account'],
            'start_time': alert['start_time'],
            'alert_status': alert['raw_alert'].get('status', 'Unknown')
        }
        enriched_alerts.append(enriched_alert)
    
    if skipped_count > 0:
        print(f"  Skipped {skipped_count} alerts with non-compliance policies")
    
    # Generate output filename
    if args.output_file:
        output_file = output_dir / args.output_file
    else:
        # Base filename
        base_filename = f"lacework_alerts_{start_date}_to_{end_date}"
        
        # Add report name to filename if specified
        if args.report:
            # Convert special characters to dashes for safe filename
            import re
            safe_report_name = re.sub(r'[():.\s]+', '-', args.report)
            # Remove multiple consecutive dashes and trim
            safe_report_name = re.sub(r'-+', '-', safe_report_name).strip('-')
            base_filename += f"_{safe_report_name}"
        
        # Output to Excel format
        output_file = output_dir / f"{base_filename}.xlsx"
    
    # Get compliance status data by default (unless skipped)
    compliance_data = []
    if not args.skip_compliance:
        print("Step 7: Retrieving compliance status...")
        
        # Get AWS accounts
        aws_accounts = get_aws_accounts(credentials)
        if not aws_accounts:
            print("No AWS accounts found. Skipping compliance status.")
        else:
            # Get compliance reports for each account
            total_accounts = len(aws_accounts)
            print(f"Processing compliance reports for {total_accounts} AWS accounts...")
            
            # Dictionary to group compliance items by policy ID
            compliance_by_policy = {}
            
            for i, account in enumerate(aws_accounts, 1):
                account_id = account['account_id']
                if not account_id:
                    continue
                
                # Show progress indicator
                progress_percent = (i / total_accounts) * 100
                print(f"[{i}/{total_accounts}] ({progress_percent:.1f}%) Getting compliance report for account: {account_id}")
                
                # Use report name from either -r or --compliance-report
                report_name = args.report or args.compliance_report
                compliance_report = get_compliance_report(
                    account_id, 
                    credentials, 
                    cache_dir, 
                    report_name
                )
                
                if compliance_report:
                    # Parse compliance data
                    parsed_compliance = parse_compliance_report_data(compliance_report, policy_details)
                    
                    # Group by policy ID
                    for item in parsed_compliance:
                        policy_id = item['policy_id']
                        if policy_id not in compliance_by_policy:
                            # Get policy details from cache (same as used for alerts)
                            policy_info = policy_details.get(policy_id, {})
                            
                            compliance_by_policy[policy_id] = {
                                'policy_id': policy_id,
                                'policy_name': policy_info.get('policy_name', item.get('policy_name', 'Unknown')),
                                'description': policy_info.get('description', item.get('description', 'N/A')),
                                'remediation': policy_info.get('remediation', item.get('remediation', 'N/A')),
                                'severity': item['severity'],
                                'resources': [],
                                'regions': set(),
                                'accounts': set(),
                                'compliance_status': item['compliance_status'],
                                'raw_data': item['raw_data']
                            }
                        
                        # Add resource information
                        compliance_by_policy[policy_id]['resources'].append(item['resource'])
                        compliance_by_policy[policy_id]['regions'].add(item['region'])
                        compliance_by_policy[policy_id]['accounts'].add(item['account'])
                    
                    print(f"  → Found {len(parsed_compliance)} compliance items")
                else:
                    print(f"  → No compliance data found")
            
            # Get policy details for any compliance policies not already cached
            print("Retrieving policy details for compliance policies...")
            compliance_policy_ids = set(compliance_by_policy.keys())
            missing_policy_ids = compliance_policy_ids - set(policy_details.keys())
            
            if missing_policy_ids:
                print(f"Found {len(missing_policy_ids)} compliance policies not in alerts cache, retrieving details...")
                for policy_id in missing_policy_ids:
                    policy_info = get_policy_details_with_retry(client, policy_id, policy_cache_dir)
                    policy_details[policy_id] = policy_info
            
            # Convert grouped data back to list format
            compliance_data = []
            for policy_id, policy_data in compliance_by_policy.items():
                # Get updated policy details
                policy_info = policy_details.get(policy_id, {})
                
                # Format resources as numbered list, sorted alphabetically
                sorted_resources = sorted(policy_data['resources'])
                if len(sorted_resources) == 1:
                    resource = sorted_resources[0]
                else:
                    resource = '\n'.join(f"{i+1}. {res}" for i, res in enumerate(sorted_resources))
                
                # Format regions and accounts
                region = '\n'.join(sorted(policy_data['regions'])) if policy_data['regions'] else 'Unknown'
                account = '\n'.join(sorted(policy_data['accounts'])) if policy_data['accounts'] else 'Unknown'
                
                # Map numeric severity to text labels
                def map_compliance_severity(severity_value):
                    """Map numeric severity from compliance reports to text labels."""
                    severity_map = {
                        '1': 'Critical',
                        '2': 'High', 
                        '3': 'Medium',
                        '4': 'Low',
                        '5': 'Info',
                        '6': 'Info',  # Additional severity levels map to Info
                        '0': 'Info',  # Some systems use 0 for informational
                        '': 'Info',   # Empty severity defaults to Info
                        None: 'Info'  # None severity defaults to Info
                    }
                    return severity_map.get(str(severity_value), 'Info')
                
                compliance_item = {
                    'policy_id': policy_data['policy_id'],
                    'policy_name': policy_info.get('policy_name', policy_data['policy_name']),
                    'description': policy_info.get('description', policy_data['description']),
                    'remediation': policy_info.get('remediation', policy_data['remediation']),
                    'severity': map_compliance_severity(policy_data['severity']),
                    'resource': resource,
                    'region': region,
                    'account': account,
                    'compliance_status': policy_data['compliance_status'],
                    'raw_data': policy_data['raw_data']
                }
                compliance_data.append(compliance_item)
            
            print(f"\nCompleted processing all accounts. Found {len(compliance_data)} unique policies with compliance issues")
    
    # Write output
    print("Step 8: Writing Excel output...")
    write_alert_excel(enriched_alerts, output_file, start_date, end_date, credentials['account'], compliance_data, args.report, args.compliance_report)
    
    # Final summary
    print("\n=== Final Summary ===")
    print(f"Date Range: {start_date} to {end_date}")
    print(f"Total alerts processed: {len(enriched_alerts)}")
    if compliance_data:
        print(f"Total compliance items: {len(compliance_data)}")
    print(f"Unique policies: {len(unique_policy_ids)}")
    print(f"Output file: {output_file}")
    
    # Alert severity distribution
    alert_severity_counts = {}
    for alert in enriched_alerts:
        severity = alert['severity'].title() if alert['severity'] != 'Unknown' else 'Unknown'
        alert_severity_counts[severity] = alert_severity_counts.get(severity, 0) + 1
    
    # Define severity order for sorting
    severity_order = {
        'Critical': 1, 'High': 2, 'Medium': 3, 'Low': 4, 'Info': 5, 'Unknown': 6
    }
    
    print("\nAlert severity distribution:")
    for severity, count in sorted(alert_severity_counts.items(), key=lambda x: severity_order.get(x[0], 999)):
        print(f"  {severity}: {count}")
    
    # Compliance severity distribution
    if compliance_data:
        compliance_severity_counts = {}
        for item in compliance_data:
            severity = item['severity']  # Already mapped to text labels
            compliance_severity_counts[severity] = compliance_severity_counts.get(severity, 0) + 1
        
        print("\nCompliance severity distribution:")
        for severity, count in sorted(compliance_severity_counts.items(), key=lambda x: severity_order.get(x[0], 999)):
            print(f"  {severity}: {count}")
    
    print(f"\nLacework reporting analysis complete!")


if __name__ == "__main__":
    main()

