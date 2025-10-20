"""
Excel report generation functionality for Lacework Alert Reporting.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from pathlib import Path


class ExcelGenerator:
    """Handles Excel report generation."""
    
    def __init__(self):
        """Initialize Excel generator."""
        self.workbook = Workbook()
        # Keep the default sheet for now, we'll remove it when we add our first sheet
    
    def create_alerts_sheet(self, alerts: List[Dict[str, Any]], sheet_name: str = "Alerts") -> None:
        """
        Create alerts sheet with the given data.
        
        Args:
            alerts: List of alert dictionaries
            sheet_name: Name of the sheet
        """
        if not alerts:
            return
        
        # Remove default sheet if this is the first sheet we're creating
        if len(self.workbook.worksheets) == 1 and self.workbook.active.title == "Sheet":
            self.workbook.remove(self.workbook.active)
        
        ws = self.workbook.create_sheet(title=sheet_name)
        
        # Define fieldnames
        fieldnames = [
            'Policy ID', 'Policy Title', 'Description', 'Remediation Steps', 'Severity',
            'Resource', 'Region', 'Account', 'Tags', 'Alert Status', 'Alert ID'
        ]
        
        # Write headers
        for col, fieldname in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col, value=fieldname)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row, alert in enumerate(alerts, 2):
            row_data = {
                'Policy ID': alert.get('policy_id', 'N/A'),
                'Policy Title': alert.get('policy_title', 'N/A'),
                'Description': alert.get('description', 'N/A'),
                'Remediation Steps': alert.get('remediation_steps', 'N/A'),
                'Severity': alert.get('severity', 'N/A'),
                'Resource': alert.get('resource', 'N/A'),
                'Region': alert.get('region', 'N/A'),
                'Account': alert.get('account', 'N/A'),
                'Tags': alert.get('tags', 'N/A'),
                'Alert Status': alert.get('alert_status', 'N/A'),
                'Alert ID': alert.get('alert_id', 'N/A')
            }
            
            for col, fieldname in enumerate(fieldnames, 1):
                ws.cell(row=row, column=col, value=row_data[fieldname])
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def create_compliance_sheet(self, compliance_data: List[Dict[str, Any]], sheet_name: str = "Compliance Status") -> None:
        """
        Create compliance status sheet with the given data.
        
        Args:
            compliance_data: List of compliance item dictionaries
            sheet_name: Name of the sheet
        """
        if not compliance_data:
            return
        
        # Remove default sheet if this is the first sheet we're creating
        if len(self.workbook.worksheets) == 1 and self.workbook.active.title == "Sheet":
            self.workbook.remove(self.workbook.active)
        
        ws = self.workbook.create_sheet(title=sheet_name)
        
        # Define fieldnames
        fieldnames = [
            'Policy ID', 'Policy Title', 'Description and Remediation', 'Severity',
            'Resource', 'Region', 'Account', 'Tags', 'Technical Owner', 
            'Business Owner', 'Environment', 'Tag Source'
        ]
        
        # Write headers with blue background and white text
        for col, fieldname in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col, value=fieldname)
            cell.font = Font(bold=True, color="FFFFFF")  # White text
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue background
            cell.alignment = Alignment(horizontal="center")
        
        # Sort compliance data by Severity, Policy Title, Account, then Resource
        def sort_key(item):
            severity_order = {'Critical': 1, 'High': 2, 'Medium': 3, 'Low': 4, 'Info': 5}
            severity = item.get('severity', 'Info')
            severity_rank = severity_order.get(severity, 6)
            
            return (
                severity_rank,
                item.get('policy_title', ''),
                item.get('account', ''),
                item.get('resource', '')
            )
        
        sorted_compliance_data = sorted(compliance_data, key=sort_key)
        
        # Write data
        for row, item in enumerate(sorted_compliance_data, 2):
            row_data = {
                'Policy ID': item.get('policy_id', 'N/A'),
                'Policy Title': item.get('policy_title', 'N/A'),
                'Description and Remediation': item.get('remediation_steps', 'N/A'),
                'Severity': item.get('severity', 'N/A'),
                'Resource': item.get('resource', 'N/A'),
                'Region': item.get('region', 'N/A'),
                'Account': item.get('account', 'N/A'),
                'Tags': item.get('tags', 'N/A'),
                'Technical Owner': item.get('technical_owner', 'N/A'),
                'Business Owner': item.get('business_owner', 'N/A'),
                'Environment': item.get('environment', 'N/A'),
                'Tag Source': item.get('tag_source', 'N/A')
            }
            
            for col, fieldname in enumerate(fieldnames, 1):
                cell = ws.cell(row=row, column=col, value=row_data[fieldname])
                
                # Make Description and Remediation links clickable
                if fieldname == 'Description and Remediation' and row_data[fieldname] != 'N/A':
                    link_value = row_data[fieldname]
                    if isinstance(link_value, str) and link_value.startswith('http'):
                        cell.hyperlink = link_value
                        cell.font = Font(color="0000FF", underline="single")
                        cell.value = link_value  # Show the actual URL
        
        # Add auto-filter to the data range
        last_row = len(sorted_compliance_data) + 1  # +1 for header row
        last_col_letter = get_column_letter(len(fieldnames))
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, ws) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set minimum width and maximum width
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def save_workbook(self, output_path: Path) -> None:
        """
        Save the workbook to the specified path.
        
        Args:
            output_path: Path where to save the Excel file
        """
        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Save the workbook
        self.workbook.save(output_path)
        print(f"Successfully wrote Excel report to {output_path}")
    
    def get_sheet_count(self) -> int:
        """Get the number of sheets in the workbook."""
        return len(self.workbook.worksheets)
    
    def get_sheet_names(self) -> List[str]:
        """Get the names of all sheets in the workbook."""
        return [sheet.title for sheet in self.workbook.worksheets]
